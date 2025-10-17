from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import get_db
from app.models.models import Usuario, Proposta, Contrato, Cronograma, Empresa, Consultor, AlocacaoCronograma
from app.auth import get_current_user

router = APIRouter()

@router.get("/pdf/{tipo}")
async def gerar_relatorio_pdf(
    tipo: str,
    data_inicial: date = Query(None),
    data_final: date = Query(None),
    status: str = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1f3a'),
        spaceAfter=30,
        alignment=1
    )
    
    if tipo == 'propostas':
        title = Paragraph("Relatório de Propostas", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        query = db.query(Proposta).join(Empresa)
        if data_inicial:
            query = query.filter(Proposta.data_proposta >= data_inicial)
        if data_final:
            query = query.filter(Proposta.data_proposta <= data_final)
        if status:
            query = query.filter(Proposta.status == status)
        
        propostas = query.all()
        
        data = [['Nº Proposta', 'Empresa', 'Valor', 'Status', 'Data']]
        for p in propostas:
            data.append([
                p.numero_proposta or '-',
                p.empresa.nome[:30] if p.empresa else '-',
                f'R$ {p.valor_proposta:,.2f}' if p.valor_proposta else '-',
                p.status or '-',
                p.data_proposta.strftime('%d/%m/%Y') if p.data_proposta else '-'
            ])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
    elif tipo == 'contratos':
        title = Paragraph("Relatório de Contratos", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        query = db.query(Contrato).join(Proposta)
        if data_inicial:
            query = query.filter(Contrato.data_assinatura >= data_inicial)
        if data_final:
            query = query.filter(Contrato.data_assinatura <= data_final)
        
        contratos = query.all()
        
        data = [['Nº Contrato', 'Valor', 'Assinatura', 'Vencimento', 'Status Pgto']]
        for c in contratos:
            data.append([
                c.numero_contrato or '-',
                f'R$ {c.valor:,.2f}' if c.valor else '-',
                c.data_assinatura.strftime('%d/%m/%Y') if c.data_assinatura else '-',
                c.data_vencimento.strftime('%d/%m/%Y') if c.data_vencimento else '-',
                c.status_pagamento or '-'
            ])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
    elif tipo == 'cronogramas':
        title = Paragraph("Relatório de Cronogramas", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        cronogramas = db.query(Cronograma).join(Proposta).all()
        
        data = [['Proposta', 'Início', 'Término', 'Hrs Prev.', '% Conclusão', 'Status']]
        for cr in cronogramas:
            data.append([
                cr.proposta.numero_proposta if cr.proposta else '-',
                cr.data_inicio.strftime('%d/%m/%Y') if cr.data_inicio else '-',
                cr.data_termino.strftime('%d/%m/%Y') if cr.data_termino else '-',
                str(cr.horas_previstas) if cr.horas_previstas else '-',
                f'{cr.percentual_conclusao}%' if cr.percentual_conclusao else '0%',
                cr.status or '-'
            ])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
    else:
        title = Paragraph("Relatório Geral do Sistema", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        total_propostas = db.query(Proposta).count()
        total_contratos = db.query(Contrato).count()
        total_empresas = db.query(Empresa).count()
        
        info_text = f"""
        <b>Total de Empresas:</b> {total_empresas}<br/>
        <b>Total de Propostas:</b> {total_propostas}<br/>
        <b>Total de Contratos:</b> {total_contratos}<br/>
        <b>Data do Relatório:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}
        """
        elements.append(Paragraph(info_text, styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{tipo}.pdf"}
    )

@router.get("/excel/{tipo}")
async def exportar_excel(
    tipo: str,
    data_inicial: date = Query(None),
    data_final: date = Query(None),
    status: str = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    if tipo == 'propostas':
        ws.title = "Propostas"
        headers = ['Nº Proposta', 'Empresa', 'CNPJ', 'Consultor', 'Solução', 'Valor', 'Status', 'Data Proposta']
        ws.append(headers)
        
        query = db.query(Proposta).join(Empresa)
        if data_inicial:
            query = query.filter(Proposta.data_proposta >= data_inicial)
        if data_final:
            query = query.filter(Proposta.data_proposta <= data_final)
        if status:
            query = query.filter(Proposta.status == status)
        
        for p in query.all():
            ws.append([
                p.numero_proposta,
                p.empresa.nome if p.empresa else '',
                p.empresa.cnpj if p.empresa else '',
                p.consultor.nome if p.consultor else '',
                p.solucao or '',
                float(p.valor_proposta) if p.valor_proposta else 0,
                p.status or '',
                p.data_proposta
            ])
    
    elif tipo == 'contratos':
        ws.title = "Contratos"
        headers = ['Nº Contrato', 'Nº Proposta', 'Valor', 'Data Assinatura', 'Data Vencimento', 'Status Pagamento']
        ws.append(headers)
        
        query = db.query(Contrato).join(Proposta)
        if data_inicial:
            query = query.filter(Contrato.data_assinatura >= data_inicial)
        if data_final:
            query = query.filter(Contrato.data_assinatura <= data_final)
        
        for c in query.all():
            ws.append([
                c.numero_contrato,
                c.proposta.numero_proposta if c.proposta else '',
                float(c.valor) if c.valor else 0,
                c.data_assinatura,
                c.data_vencimento,
                c.status_pagamento or ''
            ])
    
    elif tipo == 'cronogramas':
        ws.title = "Cronogramas"
        headers = ['Proposta', 'Data Início', 'Data Término', 'Horas Previstas', 'Horas Executadas', '% Conclusão', 'Status']
        ws.append(headers)
        
        for cr in db.query(Cronograma).join(Proposta).all():
            ws.append([
                cr.proposta.numero_proposta if cr.proposta else '',
                cr.data_inicio,
                cr.data_termino,
                float(cr.horas_previstas) if cr.horas_previstas else 0,
                float(cr.horas_executadas) if cr.horas_executadas else 0,
                float(cr.percentual_conclusao) if cr.percentual_conclusao else 0,
                cr.status or ''
            ])
    
    else:
        ws.title = "Exportacao Completa"
        
        ws.append(['PROPOSTAS'])
        ws.append(['Nº Proposta', 'Empresa', 'Valor', 'Status', 'Data'])
        for p in db.query(Proposta).join(Empresa).all():
            ws.append([
                p.numero_proposta,
                p.empresa.nome if p.empresa else '',
                float(p.valor_proposta) if p.valor_proposta else 0,
                p.status or '',
                p.data_proposta
            ])
        
        ws.append([])
        ws.append(['CONTRATOS'])
        ws.append(['Nº Contrato', 'Valor', 'Status Pagamento', 'Data Vencimento'])
        for c in db.query(Contrato).all():
            ws.append([
                c.numero_contrato,
                float(c.valor) if c.valor else 0,
                c.status_pagamento or '',
                c.data_vencimento
            ])
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=exportacao_{tipo}.xlsx"}
    )

@router.get("/cronograma-pdf")
async def exportar_cronograma_pdf(
    ano: int,
    mes: int,
    consultor_id: int = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a1f3a'),
        spaceAfter=30,
        alignment=1
    )
    
    mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes-1]
    title = Paragraph(f"Cronograma de Alocações - {mes_nome}/{ano}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    data_inicio = date(ano, mes, 1)
    import calendar
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    data_fim = date(ano, mes, ultimo_dia)
    
    query = db.query(AlocacaoCronograma).join(Consultor).filter(
        AlocacaoCronograma.data >= data_inicio,
        AlocacaoCronograma.data <= data_fim
    )
    
    if consultor_id:
        query = query.filter(AlocacaoCronograma.consultor_id == consultor_id)
    
    alocacoes = query.order_by(AlocacaoCronograma.data, Consultor.nome).all()
    
    data = [['Data', 'Consultor', 'Período', 'Projeto']]
    for a in alocacoes:
        data.append([
            a.data.strftime('%d/%m/%Y'),
            a.consultor.nome[:25] if a.consultor else '-',
            'Manhã' if a.periodo == 'M' else 'Tarde',
            a.codigo_projeto or '-'
        ])
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=cronograma_{mes_nome}_{ano}.pdf"}
    )

@router.get("/cronograma-excel")
async def exportar_cronograma_excel(
    ano: int,
    mes: int,
    consultor_id: int = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    mes_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'][mes-1]
    ws.title = f"Cronograma {mes_nome}"
    
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    
    ws.append([f'Cronograma de Alocações - {mes_nome}/{ano}'])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    
    ws.append([])
    ws.append(['Data', 'Consultor', 'Período', 'Projeto'])
    
    data_inicio = date(ano, mes, 1)
    import calendar
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    data_fim = date(ano, mes, ultimo_dia)
    
    query = db.query(AlocacaoCronograma).join(Consultor).filter(
        AlocacaoCronograma.data >= data_inicio,
        AlocacaoCronograma.data <= data_fim
    )
    
    if consultor_id:
        query = query.filter(AlocacaoCronograma.consultor_id == consultor_id)
    
    alocacoes = query.order_by(AlocacaoCronograma.data, Consultor.nome).all()
    
    for a in alocacoes:
        ws.append([
            a.data.strftime('%d/%m/%Y'),
            a.consultor.nome if a.consultor else '-',
            'Manhã' if a.periodo == 'M' else 'Tarde',
            a.codigo_projeto or '-'
        ])
    
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cronograma_{mes_nome}_{ano}.xlsx"}
    )
