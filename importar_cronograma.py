import openpyxl
from datetime import datetime, date
from sqlalchemy.orm import Session
from app.database import SessionLocal, engine, Base
from app.models.models import Consultor, AlocacaoCronograma

def limpar_dados_existentes(db: Session):
    """Remove todas as alocações do cronograma existentes"""
    db.query(AlocacaoCronograma).delete()
    db.commit()
    print("✓ Dados de alocações anteriores removidos")

def importar_cronograma(arquivo_excel: str):
    """Importa dados do cronograma da planilha Excel"""
    
    db = SessionLocal()
    
    try:
        # Limpar dados existentes
        print("Limpando dados existentes...")
        limpar_dados_existentes(db)
        
        # Carregar planilha
        print(f"Carregando planilha: {arquivo_excel}")
        wb = openpyxl.load_workbook(arquivo_excel)
        ws = wb.active
        
        # Encontrar as datas (linha 7, a partir da coluna E)
        datas = []
        for col in range(5, ws.max_column + 1):  # Coluna E = 5
            valor = ws.cell(row=7, column=col).value
            if isinstance(valor, datetime):
                datas.append(valor.date())
            elif valor is None:
                break
        
        print(f"✓ Encontradas {len(datas)} datas no cronograma")
        
        # Processar consultores (a partir da linha 13)
        consultores_criados = {}
        alocacoes_criadas = 0
        
        linha = 13
        while linha <= ws.max_row:
            nome_consultor = ws.cell(row=linha, column=2).value
            nif = ws.cell(row=linha, column=3).value
            periodo = ws.cell(row=linha, column=4).value
            
            if not nome_consultor or nome_consultor == "CONSULTORES":
                linha += 1
                continue
            
            # Criar ou buscar consultor
            consultor = None
            if nif and nif in consultores_criados:
                consultor = consultores_criados[nif]
            elif nif:
                consultor = db.query(Consultor).filter(Consultor.nif == nif).first()
                if not consultor:
                    # Criar consultor se não existir
                    email = f"{nif.lower().replace('-', '')}@sistema.com"
                    consultor = Consultor(
                        nome=nome_consultor,
                        nif=nif,
                        email=email,
                        ativo=True
                    )
                    db.add(consultor)
                    db.flush()
                    print(f"  → Consultor criado: {nome_consultor} ({nif})")
                
                consultores_criados[nif] = consultor
            
            # Processar alocações para cada data
            if consultor and periodo in ['M', 'T']:
                for idx, data in enumerate(datas):
                    col = 5 + idx  # Coluna E = 5
                    codigo_projeto = ws.cell(row=linha, column=col).value
                    
                    if codigo_projeto and str(codigo_projeto).strip():
                        # Verificar se já existe alocação
                        alocacao_existe = db.query(AlocacaoCronograma).filter(
                            AlocacaoCronograma.consultor_id == consultor.id,
                            AlocacaoCronograma.data == data,
                            AlocacaoCronograma.periodo == periodo
                        ).first()
                        
                        if not alocacao_existe:
                            alocacao = AlocacaoCronograma(
                                consultor_id=consultor.id,
                                data=data,
                                periodo=periodo,
                                codigo_projeto=str(codigo_projeto).strip(),
                                nif=nif
                            )
                            db.add(alocacao)
                            alocacoes_criadas += 1
            
            linha += 1
            
            # Commit a cada 100 linhas para performance
            if linha % 100 == 0:
                db.commit()
                print(f"  Processadas {linha} linhas...")
        
        # Commit final
        db.commit()
        print(f"\n✓ Importação concluída!")
        print(f"  - Total de consultores: {len(consultores_criados)}")
        print(f"  - Total de alocações criadas: {alocacoes_criadas}")
        
    except Exception as e:
        db.rollback()
        print(f"✗ Erro na importação: {e}")
        raise
    finally:
        db.close()

if __name__ == "__main__":
    # Criar tabelas se não existirem
    Base.metadata.create_all(bind=engine)
    
    # Importar cronograma
    arquivo = "attached_assets/crnograma principal jef_1760719792866.xlsx"
    importar_cronograma(arquivo)
