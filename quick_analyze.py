import pandas as pd
import openpyxl

files = [
    ("attached_assets/CRONOGRAMA 2.0 (4)_1760642348599.xlsx", "CRONOGRAMA"),
    ("attached_assets/Controle Geral 3.0_151015_1760642348601.xlsx", "CONTROLE GERAL"),
    ("attached_assets/Controle Geral_Considerações_1760642348602.xlsx", "CONSIDERAÇÕES")
]

for file_path, file_name in files:
    print(f"\n{'='*60}\n{file_name}\n{'='*60}")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"{sheet}: {ws.max_row} linhas x {ws.max_column} colunas")
    except Exception as e:
        print(f"Erro: {e}")
